import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def style_excel_file(input_file, output_file):
    """
    Преобразует Excel файл в красивый и читаемый формат
    """
    try:
        # Загружаем файл
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # Определяем стили
        # Цвета
        HEADER_BG_COLOR = "1F4E78"  # Темно-синий
        HEADER_TEXT_COLOR = "FFFFFF"  # Белый
        ACTIVE_STATUS_COLOR = "E2EFDA"  # Светло-зеленый для Active
        ALT_ROW_COLOR = "F2F2F2"    # Светло-серый для чередующихся строк

        # Шрифты
        header_font = Font(name='Arial', size=12, bold=True, color=HEADER_TEXT_COLOR)
        data_font = Font(name='Arial', size=10)
        phone_font = Font(name='Consolas', size=10)  # Моноширинный для телефонов
        email_font = Font(name='Arial', size=10, color="0070C0", underline='single')

        # Заливка
        header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
        active_fill = PatternFill(start_color=ACTIVE_STATUS_COLOR, end_color=ACTIVE_STATUS_COLOR, fill_type="solid")
        alt_row_fill = PatternFill(start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR, fill_type="solid")

        # Выравнивание
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Границы
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Стилизуем заголовки
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Определяем индексы колонок по заголовкам
        column_indices = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                column_indices[header] = col

        # Стилизуем данные
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = left_align
                cell.font = data_font

                # Чередование цветов строк
                if row % 2 == 0:
                    cell.fill = alt_row_fill

        # Специальное форматирование для конкретных колонок
        if 'Phone' in column_indices:
            phone_col = column_indices['Phone']
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=phone_col)
                cell.font = phone_font
                cell.alignment = center_align

        if 'Email' in column_indices:
            email_col = column_indices['Email']
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=email_col)
                if cell.value:  # Только если есть email
                    cell.font = email_font
                    cell.hyperlink = f"mailto:{cell.value}"

        if 'Status' in column_indices:
            status_col = column_indices['Status']
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=status_col)
                if cell.value and str(cell.value).upper() == 'ACTIVE':
                    cell.fill = active_fill
                    cell.font = Font(name='Arial', size=10, bold=True, color="006100")
                    cell.alignment = center_align

        # Настраиваем ширину колонок
        column_widths = {
            'BIN': 15,
            'CEO': 35,
            'Address': 50,
            'Phone': 20,
            'Email': 30,
            'Status': 12
        }

        for col_name, width in column_widths.items():
            if col_name in column_indices:
                col_letter = get_column_letter(column_indices[col_name])
                ws.column_dimensions[col_letter].width = width

        # Замораживаем заголовок
        ws.freeze_panes = "A2"

        # Добавляем фильтры
        ws.auto_filter.ref = ws.dimensions

        # Сохраняем результат
        wb.save(output_file)
        print(f"Файл успешно оформлен: {output_file}")

    except Exception as e:
        print(f"Ошибка при оформлении файла: {e}")

# Дополнительная функция для улучшения читаемости телефонов
def format_phone_numbers(input_file):
    """
    Форматирует номера телефонов в единый стиль
    """
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # Находим колонку Phone
        phone_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Phone':
                phone_col = col
                break

        if phone_col:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=phone_col)
                if cell.value:
                    # Очищаем номер от лишних символов
                    phone = str(cell.value).strip()
                    # Удаляем все нецифровые символы
                    digits = ''.join(filter(str.isdigit, phone))

                    # Форматируем в международный формат
                    if digits.startswith('7') and len(digits) == 11:
                        formatted = f"+7 ({digits[1:4]}) {digits[4:7]}-{digits[7:9]}-{digits[9:]}"
                    elif digits.startswith('870') and len(digits) == 11:
                        formatted = f"+7 ({digits[2:5]}) {digits[5:8]}-{digits[8:10]}-{digits[10:]}"
                    elif len(digits) >= 10:
                        formatted = f"+{digits}"
                    else:
                        formatted = phone

                    cell.value = formatted

        wb.save(input_file)
        print("Номера телефонов отформатированы")

    except Exception as e:
        print(f"Ошибка при форматировании телефонов: {e}")

if __name__ == "__main__":
    input_filename = "filtered_output.xlsx"
    output_filename = "beautiful_output.xlsx"

    # Сначала форматируем телефоны
    format_phone_numbers(input_filename)

    # Затем применяем стили
    style_excel_file(input_filename, output_filename)

    print("Готово! Файл красиво оформлен и готов к использованию.")
