import aiohttp
import asyncio
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dataclasses import dataclass, field
from typing import Optional, List, Dict, Any, Union
import json
from datetime import datetime
import contextlib


@dataclass
class Entity:
    """Датакласс для представления информации о компании"""

    # Основная информация
    bin: str = ""
    title_ru: str = ""
    title_kz: str = ""
    address_ru: str = ""
    address_kz: str = ""
    ceo_name: str = ""
    ceo_position: str = ""
    primary_oked: str = ""
    secondary_oked: List[str] = field(default_factory=list)
    kato_code: str = ""
    kato_description: str = ""
    registration_date: str = ""
    status: str = ""
    status_description: str = ""
    years_on_market: int = 0
    months_on_market: int = 0
    is_nds: bool = False

    # Классификаторы
    krp: str = ""
    krp_description: str = ""
    kfc: str = ""
    kfc_description: str = ""
    kse: str = ""
    kse_description: str = ""
    rnn: str = ""

    # Контактная информация
    email: str = ""
    phone: str = ""
    website: str = ""
    postal_code: str = ""
    city: str = ""
    street: str = ""

    # Налоговая информация
    total_debt_kgd: float = 0.0
    total_fine_kgd: float = 0.0
    main_debt_kgd: float = 0.0
    total_debt_egov: float = 0.0
    pension_debt: float = 0.0
    medical_debt: float = 0.0
    social_debt: float = 0.0

    # Нарушения и реестры
    violation_count: int = 0
    warning_count: int = 0
    in_inactive_registry: bool = False
    in_absent_registry: bool = False
    in_fake_registry: bool = False
    in_bankrupt_registry: bool = False
    in_invalid_registry: bool = False
    in_tax_debtor_registry: bool = False
    unreliable_samruk: bool = False
    unreliable_gz: bool = False
    was_nds: bool = False

    # Связанные компании
    filials_count: int = 0
    same_address_count: int = 0
    same_ceo_count: int = 0


def safe_extract_str(data: Any, default: str = "") -> str:
    """Безопасно извлекает строковое значение из сложной структуры"""
    if data is None:
        return default

    if isinstance(data, str):
        return data

    if isinstance(data, dict):
        # Пытаемся извлечь значение из словаря
        value = data.get('value', data)
        if isinstance(value, str):
            return value
        elif isinstance(value, dict):
            return value.get('value', default) or default
        else:
            return str(value) if value is not None else default

    return str(data) if data is not None else default


def safe_extract_list(data: Any) -> List[str]:
    """Безопасно извлекает список значений"""
    if data is None:
        return []

    if isinstance(data, list):
        return [str(item) for item in data if item is not None]

    if isinstance(data, dict) and 'value' in data:
        value = data['value']
        if isinstance(value, list):
            return [str(item) for item in value if item is not None]
        return [str(value)] if value is not None else []

    return [str(data)] if data is not None else []


def safe_get(dictionary: Optional[Dict], key: str, default: Any = None) -> Any:
    """Безопасное получение значения из словаря с проверкой на None"""
    if dictionary is None:
        return default
    return dictionary.get(key, default)


def entity_from_json(company_data: Dict[str, Any], full_info: Optional[Dict[str, Any]]) -> Entity:
    """Создает объект Entity из JSON данных с безопасным извлечением значений"""
    entity = Entity()

    # Основная информация из company_data
    entity.bin = safe_extract_str(company_data.get("bin"))

    # Основная информация из full_info
    basic_info = safe_get(full_info, "basicInfo", {}) if full_info else {}

    entity.title_ru = safe_extract_str(safe_get(basic_info, "titleRu"))
    entity.title_kz = safe_extract_str(safe_get(basic_info, "titleKz"))
    entity.address_ru = safe_extract_str(safe_get(basic_info, "addressRu"))
    entity.address_kz = safe_extract_str(safe_get(basic_info, "addressKz"))

    # Информация о руководителе
    ceo_info = safe_get(basic_info, "ceo", {})
    ceo_value = safe_get(ceo_info, "value", {})
    if isinstance(ceo_info, dict):
        entity.ceo_name = safe_extract_str(safe_get(ceo_value, "title"))
        entity.ceo_position = safe_extract_str(safe_get(ceo_value, "position"))
    else:
        entity.ceo_name = safe_extract_str(ceo_value)

    entity.primary_oked = safe_extract_str(safe_get(basic_info, "primaryOKED"))

    # Дополнительные ОКЭД
    secondary_oked = safe_get(basic_info, "secondaryOKED", [])
    entity.secondary_oked = safe_extract_list(secondary_oked)

    # KATO информация
    kato_info = safe_get(basic_info, "kato", {})
    kato_value = safe_get(kato_info, "value", {})
    entity.kato_code = safe_extract_str(safe_get(kato_value, "value"))
    entity.kato_description = safe_extract_str(safe_get(kato_value, "description"))

    entity.registration_date = safe_extract_str(safe_get(basic_info, "registrationDate"))

    # Статус компании
    status_info = safe_get(basic_info, "status", {})
    status_value = safe_get(status_info, "value", {})
    entity.status = safe_extract_str(safe_get(status_value, "value"))
    entity.status_description = safe_extract_str(safe_get(status_value, "description"))

    # Время на рынке
    on_market = safe_get(basic_info, "onMarket", {})
    entity.years_on_market = safe_get(on_market, "years", 0) or 0
    entity.months_on_market = safe_get(on_market, "months", 0) or 0

    entity.is_nds = bool(safe_get(basic_info, "isNds", False))

    # Классификаторы с безопасным доступом
    krp_info = safe_get(basic_info, "krp", {})
    krp_value = safe_get(krp_info, "value", {})
    entity.krp = safe_extract_str(safe_get(krp_value, "value"))
    entity.krp_description = safe_extract_str(safe_get(krp_value, "description"))

    kfc_info = safe_get(basic_info, "kfc", {})
    kfc_value = safe_get(kfc_info, "value", {})
    entity.kfc = safe_extract_str(safe_get(kfc_value, "value"))
    entity.kfc_description = safe_extract_str(safe_get(kfc_value, "description"))

    kse_info = safe_get(basic_info, "kse", {})
    kse_value = safe_get(kse_info, "value", {})
    entity.kse = safe_extract_str(safe_get(kse_value, "value"))
    entity.kse_description = safe_extract_str(safe_get(kse_value, "description"))

    # Контактная информация
    gos_zakup_contacts = safe_get(full_info, "gosZakupContacts", {}) if full_info else {}
    egov_contacts = safe_get(full_info, "egovContacts", {}) if full_info else {}

    # Email
    email_list = safe_get(gos_zakup_contacts, "email", []) or safe_get(egov_contacts, "email", [])
    if email_list and isinstance(email_list, list) and len(email_list) > 0:
        first_email = email_list[0]
        if isinstance(first_email, dict):
            entity.email = safe_extract_str(safe_get(first_email, "value"))
        else:
            entity.email = safe_extract_str(first_email)

    # Телефон
    phone_list = safe_get(gos_zakup_contacts, "phone", []) or safe_get(egov_contacts, "phone", [])
    if phone_list and isinstance(phone_list, list) and len(phone_list) > 0:
        first_phone = phone_list[0]
        if isinstance(first_phone, dict):
            entity.phone = safe_extract_str(safe_get(first_phone, "value"))
        else:
            entity.phone = safe_extract_str(first_phone)

    entity.postal_code = safe_extract_str(safe_get(basic_info, "postalCode"))
    entity.city = safe_extract_str(safe_get(basic_info, "cityName"))
    entity.street = safe_extract_str(safe_get(basic_info, "streetName"))

    # Налоговая информация
    debts_info = safe_get(full_info, "debtsInfo", {}) if full_info else {}
    kgd_debts = safe_get(debts_info, "kgd", {})
    egov_debts = safe_get(debts_info, "egov", {})

    entity.total_debt_kgd = float(safe_get(kgd_debts, "totalDebt", 0) or 0)
    entity.total_fine_kgd = float(safe_get(kgd_debts, "totalFine", 0) or 0)
    entity.main_debt_kgd = float(safe_get(kgd_debts, "totalMainDebt", 0) or 0)
    entity.total_debt_egov = float(safe_get(egov_debts, "totalDebt", 0) or 0)
    entity.pension_debt = float(safe_get(egov_debts, "totalPensionDebt", 0) or 0)
    entity.medical_debt = float(safe_get(egov_debts, "totalMedicalDebt", 0) or 0)
    entity.social_debt = float(safe_get(egov_debts, "totalSocialDebt", 0) or 0)

    # Нарушения и реестры
    entity.violation_count = safe_get(company_data, "reestrViolationCount", 0) or 0
    entity.warning_count = safe_get(company_data, "warningCount", 0) or 0

    reestrs_info = safe_get(full_info, "reestrs", []) if full_info else []
    for reestr in reestrs_info:
        if not isinstance(reestr, dict):
            continue

        violation = safe_get(reestr, "violation")
        description = safe_extract_str(safe_get(reestr, "description", ""))

        if violation == 0:
            entity.in_inactive_registry = True
        elif violation == 1:
            entity.in_absent_registry = True
        elif violation == 4:
            entity.in_fake_registry = True
        elif violation == 3:
            entity.in_bankrupt_registry = True
        elif violation == 5:
            entity.in_invalid_registry = True
        elif violation == 2:
            entity.in_tax_debtor_registry = True

        if "Самрук-Қазына" in description:
            entity.unreliable_samruk = True
        if "государственных закупок" in description:
            entity.unreliable_gz = True
        if "Плательщик НДС" in description:
            entity.was_nds = True

    # Связанные компании
    related_info = safe_get(full_info, "relatedCompanies", {}) if full_info else {}
    filials = safe_get(related_info, "filials", {})
    same_address = safe_get(related_info, "sameAddress", {})
    same_fio = safe_get(related_info, "sameFio", {})

    entity.filials_count = safe_get(filials, "total", 0) or 0
    entity.same_address_count = safe_get(same_address, "total", 0) or 0
    entity.same_ceo_count = safe_get(same_fio, "total", 0) or 0

    return entity


class SQLiteSaver:
    """Класс для сохранения объектов Entity в SQLite с безопасным управлением ресурсами"""

    def __init__(self, db_name: str = "companies.db"):
        self.db_name = db_name
        self.conn = None
        self.cursor = None
        self.connect()

    def connect(self) -> None:
        """Устанавливает соединение с базой данных"""
        try:
            self.conn = sqlite3.connect(self.db_name)
            self.conn.execute("PRAGMA journal_mode=WAL")  # Для лучшей производительности
            self.cursor = self.conn.cursor()
            self._create_table()
        except sqlite3.Error as e:
            print(f"Ошибка подключения к базе данных: {e}")
            raise

    def _create_table(self) -> None:
        """Создает таблицу если она не существует"""
        create_table_query = """
        CREATE TABLE IF NOT EXISTS companies (
            bin TEXT PRIMARY KEY,
            title_ru TEXT,
            title_kz TEXT,
            address_ru TEXT,
            address_kz TEXT,
            ceo_name TEXT,
            ceo_position TEXT,
            primary_oked TEXT,
            secondary_oked TEXT,
            kato_code TEXT,
            kato_description TEXT,
            registration_date TEXT,
            status TEXT,
            status_description TEXT,
            years_on_market INTEGER,
            months_on_market INTEGER,
            is_nds INTEGER,
            krp TEXT,
            krp_description TEXT,
            kfc TEXT,
            kfc_description TEXT,
            kse TEXT,
            kse_description TEXT,
            rnn TEXT,
            email TEXT,
            phone TEXT,
            website TEXT,
            postal_code TEXT,
            city TEXT,
            street TEXT,
            total_debt_kgd REAL,
            total_fine_kgd REAL,
            main_debt_kgd REAL,
            total_debt_egov REAL,
            pension_debt REAL,
            medical_debt REAL,
            social_debt REAL,
            violation_count INTEGER,
            warning_count INTEGER,
            in_inactive_registry INTEGER,
            in_absent_registry INTEGER,
            in_fake_registry INTEGER,
            in_bankrupt_registry INTEGER,
            in_invalid_registry INTEGER,
            in_tax_debtor_registry INTEGER,
            unreliable_samruk INTEGER,
            unreliable_gz INTEGER,
            was_nds INTEGER,
            filials_count INTEGER,
            same_address_count INTEGER,
            same_ceo_count INTEGER
        )
        """
        try:
            self.cursor.execute(create_table_query)
            self.conn.commit()
        except sqlite3.Error as e:
            print(f"Ошибка создания таблицы: {e}")
            raise

    def save_entity(self, entity: Entity) -> bool:
        """Сохраняет или обновляет объект Entity в базе данных"""
        if not self.conn:
            self.connect()

        query = """
        INSERT OR REPLACE INTO companies (
            bin, title_ru, title_kz, address_ru, address_kz, ceo_name, ceo_position,
            primary_oked, secondary_oked, kato_code, kato_description, registration_date,
            status, status_description, years_on_market, months_on_market, is_nds,
            krp, krp_description, kfc, kfc_description, kse, kse_description, rnn,
            email, phone, website, postal_code, city, street,
            total_debt_kgd, total_fine_kgd, main_debt_kgd, total_debt_egov, pension_debt, medical_debt, social_debt,
            violation_count, warning_count,
            in_inactive_registry, in_absent_registry, in_fake_registry, in_bankrupt_registry,
            in_invalid_registry, in_tax_debtor_registry, unreliable_samruk, unreliable_gz, was_nds,
            filials_count, same_address_count, same_ceo_count
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """

        try:
            secondary_oked_str = json.dumps(entity.secondary_oked)
            bool_to_int = lambda x: 1 if x else 0

            data = (
                entity.bin,
                entity.title_ru,
                entity.title_kz,
                entity.address_ru,
                entity.address_kz,
                entity.ceo_name,
                entity.ceo_position,
                entity.primary_oked,
                secondary_oked_str,
                entity.kato_code,
                entity.kato_description,
                entity.registration_date,
                entity.status,
                entity.status_description,
                entity.years_on_market,
                entity.months_on_market,
                bool_to_int(entity.is_nds),
                entity.krp,
                entity.krp_description,
                entity.kfc,
                entity.kfc_description,
                entity.kse,
                entity.kse_description,
                entity.rnn,
                entity.email,
                entity.phone,
                entity.website,
                entity.postal_code,
                entity.city,
                entity.street,
                entity.total_debt_kgd,
                entity.total_fine_kgd,
                entity.main_debt_kgd,
                entity.total_debt_egov,
                entity.pension_debt,
                entity.medical_debt,
                entity.social_debt,
                entity.violation_count,
                entity.warning_count,
                bool_to_int(entity.in_inactive_registry),
                bool_to_int(entity.in_absent_registry),
                bool_to_int(entity.in_fake_registry),
                bool_to_int(entity.in_bankrupt_registry),
                bool_to_int(entity.in_invalid_registry),
                bool_to_int(entity.in_tax_debtor_registry),
                bool_to_int(entity.unreliable_samruk),
                bool_to_int(entity.unreliable_gz),
                bool_to_int(entity.was_nds),
                entity.filials_count,
                entity.same_address_count,
                entity.same_ceo_count,
            )

            self.cursor.execute(query, data)
            self.conn.commit()
            return True

        except sqlite3.Error as e:
            print(f"Ошибка сохранения данных для BIN {entity.bin}: {e}")
            self.conn.rollback()
            return False
        except Exception as e:
            print(f"Неожиданная ошибка при сохранении BIN {entity.bin}: {e}")
            return False

    def close(self) -> None:
        """Безопасно закрывает соединение с БД"""
        if self.cursor:
            self.cursor.close()
        if self.conn:
            self.conn.close()

    def __enter__(self):
        """Поддержка контекстного менеджера"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Гарантированное закрытие соединения при выходе из контекста"""
        self.close()


async def get_company_full_info(
    session: aiohttp.ClientSession, bin_number: str, lang: str = "ru"
) -> Optional[Dict]:
    """Асинхронно получает полную информацию о компании по BIN"""
    url = f"https://apiba.prgapp.kz/CompanyFullInfo?id={bin_number}&lang={lang}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Referer": f"https://ba.prg.kz/750000000-almaty/{bin_number}/",
        "Accept": "application/json, text/plain, */*",
    }

    try:
        async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=30)) as response:
            response.raise_for_status()
            return await response.json()
    except aiohttp.ClientError as e:
        print(f"Ошибка сети при получении информации для BIN {bin_number}: {e}")
    except asyncio.TimeoutError:
        print(f"Таймаут при получении информации для BIN {bin_number}")
    except Exception as e:
        print(f"Неожиданная ошибка при получении информации для BIN {bin_number}: {e}")

    return None


async def process_single_company(
    session: aiohttp.ClientSession, db_saver: SQLiteSaver, company_data: Dict
) -> None:
    """Асинхронно обрабатывает одну компанию: получает данные, парсит и сохраняет в БД"""
    bin_number = company_data.get("bin", "")
    if not bin_number:
        print("Пропуск компании без BIN")
        return

    print(f"Обрабатывается BIN: {bin_number}")

    try:
        full_info = await get_company_full_info(session, bin_number, "ru")
        if not full_info:
            print(f"Пропуск BIN {bin_number} из-за ошибки запроса")
            return

        entity = entity_from_json(company_data, full_info)
        success = db_saver.save_entity(entity)
        if success:
            print(f"Сохранен BIN: {bin_number}")
        else:
            print(f"Ошибка сохранения BIN: {bin_number}")

    except Exception as e:
        print(f"Критическая ошибка при обработке BIN {bin_number}: {e}")


async def main_async_parser() -> None:
    """Основная асинхронная функция парсинга"""
    # Используем контекстный менеджер для безопасного управления ресурсами БД
    with SQLiteSaver() as db_saver:
        url = "https://apiba.prgapp.kz/GetCompanyListAsync"
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Referer": "https://ba.prg.kz/750000000-almaty/",
            "Content-Type": "application/json",
            "Accept": "application/json, text/plain, */*",
        }

        # Создаем aiohttp сессию с ограничением соединений
        connector = aiohttp.TCPConnector(limit=10, limit_per_host=5)
        async with aiohttp.ClientSession(connector=connector) as session:
            for page in range(1, 500000):  # Все так же 2 страницы для примера
                print(f"Загружаем страницу {page}...")

                data = {
                    "page": page,
                    "pageSize": 50,
                    "market": {},
                    "tax": {},
                    "krp": [],
                    "oked": [],
                    "kato": [],
                }

                try:
                    async with session.post(url, headers=headers, json=data, timeout=30) as response:
                        response.raise_for_status()
                        result = await response.json()
                        companies_data = result.get("results", [])

                    # Создаем список задач для асинхронного выполнения с ограничением
                    tasks = []
                    for company in companies_data:
                        task = asyncio.create_task(
                            process_single_company(session, db_saver, company)
                        )
                        tasks.append(task)

                    # Ограничиваем количество одновременных задач
                    batch_size = 5
                    for i in range(0, len(tasks), batch_size):
                        batch = tasks[i:i + batch_size]
                        await asyncio.gather(*batch, return_exceptions=True)
                        await asyncio.sleep(1)  # Задержка между батчами

                    print(f"Страница {page} обработана.")

                except Exception as e:
                    print(f"Ошибка на странице {page}: {e}")
                    continue

    #print("Парсинг завершен! Данные сохранены в базу companies.db")
    # Теперь можно экспортировать из БД в Excel

    #export_db_to_excel("companies.db", "companies_final2.xlsx")


def export_db_to_excel(db_name: str, excel_filename: str) -> None:
    """Экспортирует все данные из SQLite базы в Excel файл"""
    try:
        with sqlite3.connect(db_name) as conn:
            cursor = conn.cursor()
            cursor.execute("""
            SELECT * FROM companies
            WHERE phone IS NOT NULL
            AND phone != ''
            AND phone != ' '
            AND ceo_name IS NOT NULL
            AND ceo_name != ''
            AND ceo_name != ' ';
            """)
            rows = cursor.fetchall()
            column_names = [description[0] for description in cursor.description]

        wb = Workbook()
        ws = wb.active
        ws.title = "Companies Data"

        # Записываем заголовки
        for col_num, column_name in enumerate(column_names, 1):
            cell = ws.cell(row=1, column=col_num, value=column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Записываем данные
        for row_num, row in enumerate(rows, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        # Настраиваем ширину столбцов
        for col_num, column_name in enumerate(column_names, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(15, len(str(column_name)) + 2)

        wb.save(excel_filename)
        print(f"Данные экспортированы из БД в файл {excel_filename}")

    except Exception as e:
        print(f"Ошибка при экспорте в Excel: {e}")


if __name__ == "__main__":
    try:
        asyncio.run(main_async_parser())
        export_db_to_excel("companies.db", "companies_final_02.xlsx")
    except KeyboardInterrupt:
        print("\nПарсинг прерван пользователем")
    except Exception as e:
        print(f"Критическая ошибка в основном потоке: {e}")
