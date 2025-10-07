import re

import openpyxl
from openpyxl import Workbook


def validate_phone(phone_value):
    """
    Функция для валидации номера телефона
    Возвращает True если номер валиден, False в противном случае
    """
    if phone_value is None:
        return False

    # Преобразуем в строку
    phone_str = str(phone_value).strip()

    # Удаляем все нецифровые символы (кроме + в начале)
    if phone_str.startswith("+"):
        cleaned = "+" + re.sub(r"\D", "", phone_str[1:])
    else:
        cleaned = re.sub(r"\D", "", phone_str)

    # Проверяем различные форматы номеров
    patterns = [
        r"^\+7\d{10}$",  # +71234567890 (Россия)
        r"^8\d{10}$",  # 81234567890 (Россия)
        r"^\+?\d{11,15}$",  # Международные номера
        r"^\+1\d{10}$",  # +11234567890 (США/Канада)
        r"^\+44\d{9,10}$",  # +441234567890 (Великобритания)
    ]

    return any(re.match(pattern, cleaned) for pattern in patterns)


def filter_excel_by_phone(input_file, output_file, phone_column="Phone"):
    """
    Фильтрует Excel файл, оставляя только строки с валидными номерами телефонов

    Args:
        input_file (str): путь к исходному файлу
        output_file (str): путь для сохранения отфильтрованного файла
        phone_column (str): название колонки с телефонами
    """

    try:
        # Загружаем исходный файл
        workbook = openpyxl.load_workbook(input_file)
        sheet = workbook.active

        # Находим индекс колонки с телефонами
        phone_col_index = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == phone_column:
                phone_col_index = col
                break

        if phone_col_index is None:
            raise ValueError(f"Колонка '{phone_column}' не найдена в файле")

        # Создаем новый workbook для результатов
        new_workbook = Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Filtered Data"

        # Копируем заголовки
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(row=1, column=col).value = sheet.cell(
                row=1, column=col
            ).value

        # Фильтруем строки
        valid_rows_count = 1  # начинаем с 1, т.к. заголовок уже есть

        for row in range(2, sheet.max_row + 1):
            phone_value = sheet.cell(row=row, column=phone_col_index).value

            if validate_phone(phone_value):
                valid_rows_count += 1

                # Копируем всю строку
                for col in range(1, sheet.max_column + 1):
                    new_sheet.cell(row=valid_rows_count, column=col).value = sheet.cell(
                        row=row, column=col
                    ).value

        # Сохраняем результат
        new_workbook.save(output_file)
        print(
            f"Файл успешно отфильтрован! Сохранено {valid_rows_count - 1} валидных записей."
        )

    except Exception as e:
        print(f"Произошла ошибка: {e}")


# Альтернативная версия с более строгой валидацией
def validate_phone_strict(phone_value):
    """
    Более строгая валидация номера телефона
    """
    if phone_value is None:
        return False

    phone_str = str(phone_value).strip()

    # Базовые проверки
    if len(phone_str) < 10:
        return False

    # Удаляем все нецифровые символы (кроме + в начале)
    if phone_str.startswith("+"):
        digits = re.sub(r"\D", "", phone_str[1:])
        full_number = "+" + digits
    else:
        digits = re.sub(r"\D", "", phone_str)
        full_number = digits

    # Проверяем длину номера
    if len(digits) < 10 or len(digits) > 15:
        return False

    return True


# Пример использования
if __name__ == "__main__":
    # Настройки
    input_filename = "active_companies_20250911_001422.xlsx"
    output_filename = "filtered_output.xlsx"
    phone_column_name = "Phone"  # название колонки с телефонами

    # Запуск фильтрации
    filter_excel_by_phone(input_filename, output_filename, phone_column_name)
