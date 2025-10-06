from typing import List, Dict, Set, Tuple
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from search import AdataAPI
from data import Entity, SQLiteSaver

class BinMarker():
    '''Оптимизированный класс для поиска актуальных БИНов через API'''

    def __init__(self, api: AdataAPI, logger, max_workers: int = 5):
        self.api = api
        self.logger = logger
        self.found_bins: Set[str] = set()
        self.max_workers = max_workers

    def _check_ceo_parallel(self, entity: Entity) -> Tuple[str, bool]:
        """Проверка CEO через API (для многопоточного выполнения)"""
        try:
            result = self.api.search(entity.bin)
            return (entity.bin, bool(result))
        except Exception as e:
            self.logger.error(f"Error checking CEO {entity.ceo_name}: {e}")
            return (entity.bin, False)

    def mark_actual_bins(self, entities: List[Entity]) -> Set[str]:
        """Массовая проверка БИНов с использованием многопоточности"""
        self.logger.info(f"Starting bulk check for {len(entities)} entities")

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            # Создаем futures для всех entities
            futures = {
                executor.submit(self._check_ceo_parallel, entity): entity.bin
                for entity in entities
            }

            # Обрабатываем результаты по мере их поступления
            for future in as_completed(futures):
                try:
                    bin_id, is_found = future.result()
                    if is_found:
                        self.found_bins.add(bin_id)
                except Exception as e:
                    self.logger.error(f"Error processing future: {e}")

        self.logger.info(f"Found {len(self.found_bins)} active BINs")
        return self.found_bins


class XLSXGenerator:
    """Класс для генерации XLSX отчетов с использованием openpyxl"""

    def __init__(self, logger=None):
        self.logger = logger or logging.getLogger(__name__)

    def _apply_header_styles(self, worksheet):
        """Применяет стили к заголовкам"""
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def generate_xlsx_by_bins(
        self,
        bin_list: Set[str],
        all_entities: List[Entity],
        output_path: str = None
    ) -> str:
        """
        Генерирует XLSX файл только для найденных БИНов

        Args:
            bin_list: Set найденных БИНов
            all_entities: List всех entities из БД
            output_path: Путь для сохранения (опционально)

        Returns:
            Путь к созданному файлу
        """
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"active_companies_{timestamp}.xlsx"

        # Фильтруем entities только с найденными БИНами
        filtered_entities = [
            entity for entity in all_entities
            if entity.bin in bin_list
        ]

        if not filtered_entities:
            self.logger.warning("No entities found for the given BINs")
            return ""

        try:
            # Создаем рабочую книгу
            wb = Workbook()

            # Основной лист с данными
            ws_main = wb.active
            ws_main.title = "Active Companies"

            # Добавляем заголовки
            headers = [
                'BIN',
                #'Company Name',
                'CEO',
                'Address',
                'Phone',
                'Email',
                'Status',
            ]
            ws_main.append(headers)

            # Добавляем данные
            for entity in filtered_entities:
                ws_main.append([
                    entity.bin,
                    #entity.company_name,
                    entity.ceo_name,
                    entity.address_kz,
                    entity.phone,
                    entity.email,
                    'Active'
                ])

            # Применяем стили к заголовкам
            self._apply_header_styles(ws_main)

            # Настраиваем ширину колонок
            self._adjust_column_widths(ws_main)

            # Добавляем лист со статистикой
            self._add_summary_sheet(wb, len(filtered_entities), len(bin_list))

            # Сохраняем файл
            wb.save(output_path)
            self.logger.info(f"XLSX generated successfully: {output_path}")
            return output_path

        except Exception as e:
            self.logger.error(f"Error generating XLSX: {e}")
            raise

    def _adjust_column_widths(self, worksheet):
        """Автоматически настраивает ширину колонок"""
        column_widths = {}

        for row in worksheet.iter_rows():
            for i, cell in enumerate(row):
                if cell.value:
                    length = len(str(cell.value))
                    if i not in column_widths or length > column_widths[i]:
                        column_widths[i] = min(length + 2, 50)  # Максимум 50 символов

        for i, width in column_widths.items():
            column_letter = chr(65 + i)  # A, B, C, ...
            worksheet.column_dimensions[column_letter].width = width

    def _add_summary_sheet(self, workbook, total_companies: int, unique_bins: int):
        """Добавляет лист со статистикой"""
        ws_summary = workbook.create_sheet("Summary")

        # Заголовок
        ws_summary['A1'] = "Report Summary"
        ws_summary['A1'].font = Font(bold=True, size=14)

        # Данные
        summary_data = [
            ["Metric", "Value"],
            ["Total Companies Found", total_companies],
            ["Unique BINs", unique_bins],
            ["Generation Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Data Source", "Adata.kz API + Local Database"]
        ]

        for row_idx, row_data in enumerate(summary_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Заголовки таблицы
                    cell.font = Font(bold=True)

        # Применяем стили к заголовкам
        self._apply_header_styles(ws_summary)

        # Настраиваем ширину колонок
        self._adjust_column_widths(ws_summary)


def main():
    """Оптимизированная основная логика"""
    # Инициализация компонентов
    db = SQLiteSaver()
    api = AdataAPI()
    marker = BinMarker(api=api, logger=api.logger, max_workers=10)
    xlsx_gen = XLSXGenerator(logger=api.logger)

    try:
        # Получаем все данные из БД
        all_entities: List[Entity] = db.all()

        if not all_entities:
            api.logger.warning("No entities found in database")
            return

        # Массовая проверка БИНов
        found_bins = marker.mark_actual_bins(all_entities)

        if found_bins:
            # Генерация отчета
            xlsx_path = xlsx_gen.generate_xlsx_by_bins(found_bins, all_entities)
            api.logger.info(f"Report generated: {xlsx_path}")
        else:
            api.logger.info("No active companies found")

    except Exception as e:
        api.logger.error(f"Error in main process: {e}")
        raise


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO)
    main()
